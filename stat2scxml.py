#!/usr/bin/env python3


import os
import openpyxl
import time
import datetime
from sys import argv
from optparse import OptionParser
import json
from  math import  cos


def FindRowBlock(Lst, rmin, rmax, col, word):
    first = 1
    (re, rb) = (0, 0)
    for s in range(rmin, rmax + 1):
        if Lst[s][col].value == word:
            if first:
                rb = s; re = s
                first = 0
            else:
                re = s

    return rb, re


def FindRow(Lst, rmin, rmax, col, word):
    # Lst = xlsx-List
    # rmin, rmax - List row
    # col = column of str
    # str - seek row

    rnum = 0
    for s in range(rmin, rmax + 1):
        #  D = Lst[s][col].value
        # print("debug:  s = " + str(s) + str(D) )
        if Lst[s][col].value == word:
            rnum = s
            break

    if rnum == 0:

        print("Not found search_line = " + str(word) + ",  numRow = 0  \n")
        exit()

    return rnum


def PoZ(npoz, pozx):
    # converts poles or zeros from SMP-format to xml_sc3-format
    # npoz - number of poles or zeros
    # pozx - excel-cell of poles or zeros

    P = []   # pole or zero array
    pozx = pozx.replace(" ", "")  # delete all spaces
    pozx = pozx.replace("i", "j")  # change i to j
    P = pozx.split(",")
    pstr = ""
    if npoz != len(P):
        #print("Number of poles or zeroz != len(P) " + str(len(P)) + "\n")
        exit(-1)

    for i in range(0, npoz):
        RePoz = complex(P[i]).real
        ImPoz = complex(P[i]).imag
        pxl = '(' + str(RePoz) + ',' + str(ImPoz) + ')'
        pstr += pxl + " "

    pstr = pstr.rstrip()
    return pstr

# -------------------------------------------------------------------------------------------


def main():
    usage = "usage: %prog -d 'directory for scxml file' -o 'filename of scxml file' -s 'json status file' -n 'network name' -a 'array name'"
    parser = OptionParser(usage)
    parser.add_option("-d", "--dir", dest="path",
                      help="PATH for SC XML file")
    parser.add_option("-o", "--outfile", dest="outfile",
                      help="FILEMAME for SC XML file")
    parser.add_option("-s", "--status", dest="statfile",
                      help="STATUS file in json format")
    parser.add_option("-n", "--network", dest="network",
                      help="NETWORK for SC XML")
    parser.add_option("-a", "--array", dest="array",
                      help="ARRAY for SC XML")

    (options, args) = parser.parse_args()

    process(options.path, options.outfile, options.statfile, options.network, options.array)

def process(Dir_Path, scxml, statfile, network, array):

    #print(Dir_Path)
    #print(scxml)
    #print(statfile)
    #print(network)

    #print(array)

    global GRP, NetW, dT
    global workbook, SC
    global Id_Response, Id_Sensor, Id_DLogger, Id_Stream
    global T0_Response, T0_Sensor, T0_DLogger, T0_SLoc, T_SLoc, T0_STA, T0_Stream, T_Strm

    # (Prog, STA, NetW, DFile, start_t) = argv
    NetW = network
    #print(NetW)
    GRP = array   #- group name
    nsta = 0
    with open(statfile, "r") as infile:
        json_object = json.load(infile)
        #print(json_object)
        #print(len(json_object))
        for key in json_object:
            if key.upper() == array:
                #print(key, json_object[key])
                lat_dict = {}
                lon_dict = {}
                dnorth_dict = {}
                deast_dict = {}
                ref_sta = {}
                net_dict = {}
                net_dict[array] = network
                sta_dict = {}
                location_dict = {}
                location_dict[array] = "00"
                chan_dict = {}
                chan_dict[array] = "BDF"
                smprate_dict = {}
                smprate_dict[array] = "100"
                lat_sum = 0
                lon_sum = 0
                stanames = []
                stanames.append(array)
                nsta += 1
                for key2 in json_object[key]:
                    if key2 != "display":
                        #print(key2, json_object[key][key2])

                        staname = ""
                        lat = ""
                        lon = ""

                        for key3 in json_object[key][key2]:
                           #print(len(json_object[key][key2][key3]))
                            if len(json_object[key][key2]) == 6:
                                if key3 == 'name':
                                    staname = json_object[key][key2]['name']
                                    #print(staname)
                                    stanames.append(staname)
                                    nsta += 1
                                elif key3 == 'lat':
                                    lat = json_object[key][key2]['lat']
                                    lat_dict[staname] = lat
                                    lat_sum += lat
                                    #print(lat)
                                elif key3 == 'lon':
                                    lon = json_object[key][key2]['lon']
                                    lon_dict[staname] = lon
                                    lon_sum += lon
                                    #print(lon)
                                elif key3 == 'smprate':
                                    smprate_dict[staname] = json_object[key][key2]['smprate']
                                    #print(staname, smprate_dict[staname])
                                elif key3 == 'chan':
                                    chan = json_object[key][key2]['chan']
                                    #print(chan)
                                    chan_splitted = chan.split('.')
                                    net_dict[staname] = chan_splitted[0]
                                    sta_dict[staname] = chan_splitted[1]
                                    location_dict[staname] = chan_splitted[2]
                                    chan_dict[staname] = chan_splitted[3]
                                    #print(chan_splitted)
                lat_dict[array] = lat_sum/4
                lon_dict[array] = lon_sum/4
                ref_station = stanames[1]
                for n in range(len(stanames)):
                    ref_sta[stanames[n]] = ref_station
                    dnorth_dict[stanames[n]] = (lat_dict[stanames[n]] - lat_dict[ref_station])/(1/110.574)
                    deast_dict[stanames[n]] = (lon_dict[stanames[n]] - lon_dict[ref_station])/(1/111.320*cos(lat_dict[stanames[n]]))
                    #print(ref_sta[stanames[n]], stanames[n], lat_dict[ref_station], lat_dict[stanames[n]], lon_dict[stanames[n]],\
                    #      dnorth_dict[stanames[n]], deast_dict[stanames[n]] )

    n_sta = nsta
    #print(n_sta)# number of ststions
    #print(stanames)
    #for i in range(n_sta):
    #    print(stanames[i])
    #return
    #NetW = "XX"   # - station's network
    #DFile = "XX-SMP-Inventory.xlsx"  # - input xlsx-file

    #Dir_Path = "."  # here is input file
    #os.chdir(Dir_Path)
    #print("Current dir:", os.getcwd())

    #XlsFile = Dir_Path + "/" + DFile

    # create full name (with path) of output xml-file
    now = datetime.datetime.today()
    Date = now.strftime("%Y%m%d")
    Id = now.strftime("%Y%m%d%H%M%S")

    #workbook = openpyxl.load_workbook(XlsFile, data_only=True)  # open the workbook

    #ListG = workbook['CurGroup']
    #ListS = workbook['Sensor']
    #ListD = workbook['DataLogger']
    #ListC = workbook['Channel']
    #G_date = ListG[3][4].value
    G_date = now.strftime("%Y-%m-%dT%H:%M:%S")
    #print(G_date)
    #NetW = ListC[3][1].value
    #print(str(ListG))
    #return

    (SeisComp, Inventory, Network, Station, SLocation) = (1, 1, 1, 1, 1)
    if scxml:
        SC3_xml_file = Dir_Path + "/" + scxml
    else:
        SC3_xml_file = Dir_Path + "/" + NetW + "_" + GRP + "_" + Date + "_SC.xml"
    #print(SC3_xml_file)
    #return
    T0_StaGroup = 0
    T0_Sta = 10
    T0_SLoc = T0_Sta + 10
    T0_Response = T0_SLoc + 10
    T0_Sensor = T0_Response + 10   # 31
    T0_DLogger = T0_Sensor + 10    # 41
    T0_Stream = T0_DLogger + 10    # 51
    T_Strm = T0_Stream

    StaName = []; Lat = []; Lon = []; H = []; StaId = []; SlocId = []; StreamId = []
    NetName = []; LocName = [];  ChanName = []; SmpRate = []; DNorth = []; DEast = []
    RefSta = []
    for i in range(0, n_sta):
        j = i+3
        #aa = ListG[j][3].value
        StaName.append(stanames[i])
       # print("stanames[%d] = %s" % (i,stanames[i]))
        Lat.append(lat_dict[stanames[i]])
        Lon.append(lon_dict[stanames[i]])
        RefSta.append(ref_sta[stanames[i]])
        DNorth.append(dnorth_dict[stanames[i]])
        DEast.append(deast_dict[stanames[i]])
        SmpRate.append(smprate_dict[stanames[i]])
        NetName.append(net_dict[stanames[i]])
        LocName.append(location_dict[stanames[i]])
        LocName.append(location_dict[stanames[i]])
        ChanName.append(chan_dict[stanames[i]])
        H.append("0")
        #if stanames[i] != array:
        StaId.append("Station/" + stanames[i] + "/" + Id + "." + str(i + T0_Sta))
        SlocId.append("SensorLocation/" + Id + "." + str(i + T0_SLoc))
        StreamId.append("Stream#" + Id + "." + str(i + T0_Stream))
    #print(StaName)
    #print(Lat)
    #print(Lon)
    #print(H)
    #print(StaId)
    #print(SlocId)
    #print(StreamId)
    #return
    SC = open(SC3_xml_file, "w")



    SC.write('<?xml version=\"1.0\" encoding=\"UTF-8\"?>\n')
    while SeisComp:
        SC.write('<seiscomp xmlns=\"http://geofon.gfz-potsdam.de/ns/seiscomp3-schema/0.10\" version=\"0.10\">\n')

        while Inventory:
            SC.write('  <Inventory>\n')
            #n = T0_Sta
            #StaId.append("Station/" + Id + "." + str(n))

            SC.write("\t<stationGroup publicID=\"" + "StationGroup/" + Id + "."+str(T0_StaGroup) +
                     "\" code =\"" + GRP + "\">\n")
            SC.write("\t  <type>array</type>\n")
            SC.write("\t  <description>" + GRP + " group " + NetW + "</description>\n")

            for i in range(0, n_sta):
                #print(stanames[i])
                if stanames[i] != array:
                    SC.write("\t  <stationReference>" + str(StaId[i]) + "</stationReference>\n")
                #n = n + 1
                #StaId.append("Station/" + Id + "." + str(n))
                #print('StaId = %s \n' % StaId[i])

            SC.write("\t</stationGroup>\n")

            # Sensor start
            #C = []
            SensorName = "Ermak"
            #(col_min, col_max) = (ListS.min_column, ListS.max_column)

            #for j in range(col_min - 1, col_max):
            #C.append(SensorName)
            nSensor = 1
            SensorId = "Sensor/" + Id + "." + str(T0_Sensor)
            ResponseId = "ResponsePAZ/" + Id + "." + str(T0_Response)
            #SC.write('\t' + "<sensor publicID=\"" + SensorId + "\" name=\"" + SensorName + "\" response=\"" +
            #         ResponseId + "\">\n")
            SC.write('\t' + "<sensor publicID=\"" + SensorId + "\" name=\"" + SensorName + "\">\n")
            SC.write('\t' + "  <description>" + "YYYY pressure microbarometer" + "</description>" + '\n')
            SC.write('\t' + "  <model>" + "YYYY" + "</model>" + '\n')
            SC.write('\t' + "  <lowFrequency>" + "0.05" + "</lowFrequency>" + '\n')
            SC.write('\t' + "  <highFrequency>" + "20" + "</highFrequency>" + '\n')
            SC.write('\t' + "</sensor>" + '\n')
            #print('There are created  = %d = Sensors\n' % nSensor)
            # Sensor end

            # DataLogger start
            C = []
            DLoggerName = "ERMAK"
            #(col_min, col_max) = (ListD.min_column, ListD.max_column)

            #for j in range(col_min - 1, col_max):
            #    C.append(ListS[3][j].value)

            DLogger = 1
            smprate = 100
            DLoggerId = "Datalogger/" + Id + "." + str(T0_DLogger)

            SC.write('\t' + "<datalogger publicID=\"" + DLoggerId + "\" name=\"" + DLoggerName + "\">\n")
            SC.write('\t' + "  <gain>1</gain>" + '\n')
            SC.write('\t' + "  <decimation sampleRateNumerator=\"" + str(smprate) + "\" sampleRateDenominator=\"1\"></decimation>\n")
            SC.write('\t' + "</datalogger>" + "\n")
            #print('There are created  = %d = DataLoggers\n' % nDLogger)

            # DataLogger end


            # ResponsePAZ start
            #C = []
            #(col_min, col_max) = (ListS.min_column, ListS.max_column)

            #for j in range(col_min - 1, col_max):
            #    C.append(ListS[3][j].value)
            #nResponse = 1

            #poles = PoZ(C[14], C[16])
            #print("Poles: " + str(poles))
            #if C[13] != 0:
                #zeros = PoZ(C[13], C[15])
                #print("Zeros: " + str(zeros) + '\n')
            #else:
                #nZer = 0
                #print("Zeros: " + f"{C[13]}" + '\n')
            SC.write("\t<responsePAZ publicID=\"" + ResponseId + "\" name=\"" + ResponseId + "\">\n")
            SC.write('\t' + "  <type>A</type>" + '\n')
            SC.write('\t' + "  <gain>1</gain>" + '\n')
            SC.write('\t' + "  <gainFrequency>1</gainFrequency>" + '\n')
            SC.write('\t' + "  <normalizationFactor>1</normalizationFactor>" + '\n')
            SC.write('\t' + "  <normalizationFrequency>1</normalizationFrequency>" + '\n')
            """
            SC.write('\t' + "  <numberOfZeros>" + f"{C[13]}" + "</numberOfZeros>" + '\n')
            SC.write('\t' + "  <numberOfPoles>" + f"{C[14]}" + "</numberOfPoles>" + '\n')
            if C[13] == 0:
                SC.write('\t' + "  <zeros></zeros>" + '\n')
            else:
                SC.write('\t' + "  <zeros>" + str(zeros) + "</zeros>" + '\n')
            SC.write('\t' + "  <poles>" + str(poles) + "</poles>" + '\n')
            """
            SC.write("\t</responsePAZ>\n")
            #print('There are created  = %d = Responses\n' % nResponse)

            # ResponsePAZ end
            #C = []
            #(col_min, col_max) = (ListC.min_column, ListC.max_column)
            #for j in range(col_min - 1, col_max):
            #    C.append(ListC[3][j].value)
            
            # Network
            while Network:
                NetworkId = "Network/" + Id + "." + str(100)
#                SC.write("\t<network publicID=\"" + NetworkId + "\" code=\"" + f"{C[1]}" + "\">\n")
                SC.write("\t<network publicID=\"" + NetworkId + "\" code=\"" + NetW  + "\">\n")
                SC.write("\t  <start>" + G_date + ".0000Z</start>\n")
                SC.write("\t  <description>" + NetW + " network</description>\n")

                for i in range(0, n_sta):
                    SC.write("\t  <station publicID=\"" + StaId[i] + "\" code=\"" + StaName[i] + "\">\n")
                    SC.write("\t\t<start>" + G_date + ".0000Z</start>\n")
                    SC.write("\t\t<latitude>" + str(Lat[i]) + "</latitude>\n")
                    SC.write("\t\t<longitude>" + str(Lon[i]) + "</longitude>\n")
                    SC.write("\t\t<elevation>" + str(H[i]) + "</elevation>\n")
                    if StaName[i] ==  array:
                        SC.write("\t\t<type>ar</type>\n")
                        SC.write("\t\t<remark>dnorth:0,deast:0,refsta:" + RefSta[i] + "</remark>\n")
                    else:
                        SC.write("\t\t<type>ss</type>\n")
                        SC.write("\t\t<remark>dnorth:" + str(DNorth[i]) + ",deast:" + str(DEast[i]) + ",refsta:" + RefSta[i] + "</remark>\n")
                    SC.write("\t\t<sensorLocation publicID=\"" + SlocId[i] + "\" code=\"" + str(LocName[i]) + "\">\n")
                    SC.write("\t\t  <start>" + G_date + ".0000Z</start>\n")
                    SC.write("\t\t  <latitude>" + str(Lat[i]) + "</latitude>\n")
                    SC.write("\t\t  <longitude>" + str(Lon[i]) + "</longitude>\n")
                    SC.write("\t\t  <elevation>" + str(H[i]) + "</elevation>\n")
                    if StaName[i] ==  array:
                        SC.write("\t\t  <stream code=\"" + str(ChanName[i]) + "\">\n")
                    else:
                        SC.write("\t\t  <stream publicID=\"" + StreamId[i] + "\" code=\"" + str(ChanName[i]) + \
                             "\" sensor=\"" + SensorId + "\">\n")
                    SC.write("\t\t\t<start>" + G_date + ".0000Z</start>\n")

                    SC.write("\t\t\t<sampleRateNumerator>" + SmpRate[i] + "</sampleRateNumerator>\n")
                    SC.write("\t\t\t<sampleRateDenominator>1</sampleRateDenominator>\n")
                    SC.write("\t\t\t<depth>0</depth>\n")
                    SC.write("\t\t\t<azimuth>0</azimuth>\n")
                    SC.write("\t\t\t<dip>0</dip>\n")

                    SC.write("\t\t\t<gain>" + f"1"  "</gain>\n")
                    SC.write("\t\t\t<gainFrequency>" + f"1" + "</gainFrequency>\n")
                    # print('There are created  = %d = Streams for SLocation N = %d \n' % (NStrSL, m))
                    #print('gainChan = %.1f  gainFrequency = %.1f ' % (C[20], C[19]))
#                    SC.write("\t\t\t<gainUnit>mcP</gainUnit>\n")
                    SC.write("\t\t  </stream>\n")

                    SC.write("\t\t</sensorLocation>\n")
                    Station = 0
                    SC.write("\t  </station>\n")

                Network = 0

            SC.write("\t</network>\n")
            Inventory = 0
        SC.write("  </Inventory>\n")
        SeisComp = 0

    SC.write("</seiscomp>\n")
    SC.close()

    return


if __name__ == "__main__":
    main()
